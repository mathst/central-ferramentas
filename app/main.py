"""
app/main.py — Backend FastAPI para Central de Ferramentas.

Rotas:
  GET  /           → serve index.html
  POST /gerar      → processa empresas em paralelo, retorna arquivo
  GET  /progresso  → SSE com updates por empresa durante processamento
  GET  /ping       → healthcheck para o launcher
"""
import asyncio
import io
import json
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path
from typing import AsyncGenerator

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import sys
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).parent.parent))
from processar_saldos_final import conectar_banco, processar_empresa
from processar_distrato import (
    SQL_DISTRATO, SQL_RENDIMENTO_IR, consolidar,
    TIPO_PROC_DISTRATO, COLUNAS_SAIDA,
    HEADER_FILL, HEADER_FONT, _cor_linha,
)


def _df_to_xlsx_bytes(df: pd.DataFrame, empresa: int, ano: int) -> bytes:
    """Converte DataFrame para bytes xlsx (sem salvar em disco)."""
    fmt_brl = '#,##0.00'
    df = df.copy()
    idx_mon = {
        i for i, c in enumerate(df.columns)
        if c == "Saldo Anterior" or c.startswith("Saldo ")
    }
    for i in idx_mon:
        df.iloc[:, i] = pd.to_numeric(df.iloc[:, i], errors="coerce")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        wb = writer.book
        ws = wb.create_sheet(title="Balancete")
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        writer.sheets["Balancete"] = ws

        for col_i, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_i, value=col_name)

        for row_i, tup in enumerate(df.itertuples(index=False), 2):
            for col_i, val in enumerate(tup, 1):
                if isinstance(val, float) and pd.isna(val):
                    val = None
                cell = ws.cell(row=row_i, column=col_i, value=val)
                if (col_i - 1) in idx_mon and val is not None:
                    cell.number_format = fmt_brl

        for i, col in enumerate(df.columns, 1):
            max_len = max(
                df.iloc[:, i - 1].astype(str).apply(len).max(),
                len(str(col))
            ) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max_len, 45)

        ws.freeze_panes = "A2"

    return buf.getvalue()

app = FastAPI(docs_url=None, redoc_url=None)

# ── Static files ──────────────────────────────────────────────────────────────
STATIC_DIR = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# ── Fila de progresso por request (SSE) ──────────────────────────────────────
_progress_queues: dict[str, asyncio.Queue] = {}


# ── Schemas ───────────────────────────────────────────────────────────────────

class GerarRequest(BaseModel):
    ano: int
    empresas: list[int]
    request_id: str


class DistratoRequest(BaseModel):
    empresa: int
    inicio: str  # YYYY-MM-DD
    fim: str     # YYYY-MM-DD
    request_id: str


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/ping")
async def ping():
    return {"ok": True}


@app.get("/", response_class=HTMLResponse)
async def index():
    html = (STATIC_DIR / "index.html").read_text(encoding="utf-8")
    return HTMLResponse(content=html)


@app.get("/progresso/{request_id}")
async def progresso(request_id: str, request: Request):
    """SSE stream de progresso para um request específico."""
    queue: asyncio.Queue = asyncio.Queue()
    _progress_queues[request_id] = queue

    async def event_stream() -> AsyncGenerator[bytes, None]:
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    msg = await asyncio.wait_for(queue.get(), timeout=60.0)
                except asyncio.TimeoutError:
                    yield b"data: {\"type\": \"ping\"}\n\n"
                    continue
                yield f"data: {json.dumps(msg)}\n\n".encode()
                if msg.get("type") == "done" or msg.get("type") == "error":
                    break
        finally:
            _progress_queues.pop(request_id, None)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/gerar")
async def gerar(body: GerarRequest):
    """
    Processa as empresas em paralelo (ThreadPoolExecutor, max 4 workers).
    Emite progresso via SSE na fila do request_id.
    Retorna o arquivo xlsx/zip diretamente na resposta.
    """
    ano = body.ano
    empresas = body.empresas
    request_id = body.request_id

    if not (1900 <= ano <= date.today().year + 1):
        raise HTTPException(status_code=400, detail=f"Ano inválido: {ano}")
    if not empresas:
        raise HTTPException(status_code=400, detail="Informe ao menos uma empresa.")

    queue = _progress_queues.get(request_id)

    async def push(msg: dict):
        if queue:
            await queue.put(msg)

    # Testa conectividade antes de despachar threads
    await push({"type": "step", "msg": "Conectando ao banco de dados..."})
    try:
        _test_conn = conectar_banco()
        _test_conn.close()
    except Exception as _conn_err:
        msg = f"Falha ao conectar no banco: {_conn_err}"
        await push({"type": "error", "msg": msg})
        raise HTTPException(status_code=500, detail=msg)

    buffers: list[tuple[int, bytes]] = []
    erros: list[str] = []
    total = len(empresas)
    concluidas = 0

    # Cada thread abre sua própria conexão — pyodbc não é thread-safe
    def _processar(emp: int) -> tuple[int, bytes | None, str | None]:
        try:
            conn = conectar_banco()
            try:
                df = processar_empresa(conn, emp, ano)
            finally:
                conn.close()
            if df.empty:
                return emp, None, f"Empresa {emp}: sem dados para {ano}"
            return emp, _df_to_xlsx_bytes(df, emp, ano), None
        except Exception as exc:
            return emp, None, f"Empresa {emp}: {exc}"

    await push({"type": "step", "msg": f"Processando {total} empresa(s)..."})

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=min(4, total)) as pool:
        # run_in_executor não bloqueia o event loop — evita timeout do browser
        async_futures = [
            loop.run_in_executor(pool, _processar, emp)
            for emp in empresas
        ]
        for coro in asyncio.as_completed(async_futures):
            emp_id, xlsx_bytes, erro = await coro
            concluidas += 1
            if erro:
                erros.append(erro)
                await push({"type": "progress", "msg": f"⚠ Empresa {emp_id} ignorada", "done": concluidas, "total": total})
            else:
                buffers.append((emp_id, xlsx_bytes))
                await push({"type": "progress", "msg": f"✓ Empresa {emp_id} concluída", "done": concluidas, "total": total})

    if not buffers:
        msg = "Nenhum dado gerado. " + " | ".join(erros)
        await push({"type": "error", "msg": msg})
        raise HTTPException(status_code=422, detail=msg)

    await push({"type": "step", "msg": "Montando arquivo para download..."})

    if len(buffers) == 1:
        emp_id, data = buffers[0]
        filename = f"BALANCETE_{ano}_EMP_{emp_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for emp_id, xlsx_data in sorted(buffers):
                zf.writestr(f"BALANCETE_{ano}_EMP_{emp_id}.xlsx", xlsx_data)
        data = zip_buf.getvalue()
        filename = f"BALANCETE_{ano}_LOTE.zip"
        media_type = "application/zip"

    ok_emps = ", ".join(str(e) for e, _ in sorted(buffers))
    success = f"✅ {len(buffers)} arquivo(s) gerado(s): empresa(s) {ok_emps}"
    if erros:
        success += f" | ⚠ Ignorados: {'; '.join(erros)}"
    await push({"type": "done", "msg": success, "filename": filename})

    return Response(
        content=data,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _df_distrato_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """Converte DataFrame de distrato para bytes xlsx com cores por tipo."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
        df.to_excel(writer, sheet_name="Distrato_Rendimento", index=False)
    from openpyxl import load_workbook
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    tipo_col_idx = COLUNAS_SAIDA.index("Tipo") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tipo_val = row[tipo_col_idx - 1].value
        fill = _cor_linha(str(tipo_val) if tipo_val else "")
        if fill:
            for cell in row:
                cell.fill = fill
    for col_idx, col_name in enumerate(COLUNAS_SAIDA, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@app.post("/gerar-distrato")
async def gerar_distrato(body: DistratoRequest):
    """
    Gera XLS de distrato + rendimento/IR para uma empresa e período.
    Emite progresso via SSE na fila do request_id.
    """
    import re as _re
    empresa    = body.empresa
    inicio     = body.inicio
    fim        = body.fim
    request_id = body.request_id

    if not _re.match(r"^\d{4}-\d{2}-\d{2}$", inicio) or not _re.match(r"^\d{4}-\d{2}-\d{2}$", fim):
        raise HTTPException(status_code=400, detail="Datas devem estar no formato YYYY-MM-DD.")
    if empresa <= 0:
        raise HTTPException(status_code=400, detail="Empresa inválida.")

    queue = _progress_queues.get(request_id)

    async def push(msg: dict):
        if queue:
            await queue.put(msg)

    await push({"type": "step", "msg": "Conectando ao banco de dados..."})
    try:
        _c = conectar_banco()
        _c.close()
    except Exception as _err:
        msg = f"Falha ao conectar no banco: {_err}"
        await push({"type": "error", "msg": msg})
        raise HTTPException(status_code=500, detail=msg)

    def _buscar() -> tuple[pd.DataFrame, pd.DataFrame]:
        conn = conectar_banco()
        try:
            df_dis = pd.read_sql(
                SQL_DISTRATO, conn,
                params=[empresa, TIPO_PROC_DISTRATO, inicio, fim],
            )
            df_ren = pd.read_sql(
                SQL_RENDIMENTO_IR, conn,
                params=[empresa, inicio, fim],
            )
        finally:
            conn.close()
        return df_dis, df_ren

    await push({"type": "step", "msg": f"Buscando dados — empresa {empresa}..."})

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        df_dis, df_ren = await loop.run_in_executor(pool, _buscar)

    df_final = consolidar(df_dis, df_ren)

    if df_final.empty:
        msg = "Nenhum dado encontrado para os filtros informados."
        await push({"type": "error", "msg": msg})
        raise HTTPException(status_code=422, detail=msg)

    await push({"type": "step", "msg": "Montando arquivo para download..."})

    xlsx_bytes = _df_distrato_to_xlsx_bytes(df_final)
    filename   = f"DISTRATO_EMP{empresa}_{inicio}_{fim}.xlsx"

    await push({"type": "done",
                "msg": f"✅ {len(df_final)} linha(s) geradas — empresa {empresa}",
                "filename": filename})

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="127.0.0.1", port=8000, reload=False)

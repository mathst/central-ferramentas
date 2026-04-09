"""
processar_distrato.py
Gera XLS com colunas: Empresa, Descrição, Obra, Data, Valor, CAP, Banco, Conta
Fontes:
  1. Distrato pago       — VwDesembolso WHERE TipoProc_Des = 'DES0049'
  2. Rendimento aplicação — EntSaiEmpAplic WHERE EntSai_es = 0 (receitas bancárias)
  3. IR retido           — EntSaiEmpAplic WHERE EntSai_es = 44 (despesas bancárias)
"""

import argparse
import os
import sys
from datetime import date
from pathlib import Path

import pandas as pd
import pyodbc
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Config ────────────────────────────────────────────────────────────────────

def _get_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path.cwd()


load_dotenv(_get_base_dir() / ".env")


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Gera XLS de Distrato / Rendimento / IR a partir do UAU."
    )
    # Parâmetros de negócio
    p.add_argument("--empresas", type=int, nargs="+",
                   default=[int(e) for e in os.getenv("EMPRESAS", "12").split(",")],
                   metavar="N", help="Uma ou mais empresas: --empresas 12 5 8")
    p.add_argument("--inicio",   default=os.getenv("DATA_INICIO", "2026-01-01"),
                   metavar="YYYY-MM-DD", help="Data início do período")
    p.add_argument("--fim",      default=os.getenv("DATA_FIM", date.today().isoformat()),
                   metavar="YYYY-MM-DD", help="Data fim do período (padrão: hoje)")
    p.add_argument("--tipo-proc", type=int, default=int(os.getenv("TIPO_PROC", "17")),
                   help="TipoProc_Des (tinyint) para distrato (padrão: 17)")
    p.add_argument("--status", type=int, nargs="*", default=None,
                   metavar="N",
                   help="Filtrar StatusParc_Des (0=aberto 1=parcial 2=pago). "
                        "Sem --status: traz todos. Ex: --status 2")
    # Diagnóstico: lista EntSai_es disponíveis sem gerar XLS
    p.add_argument("--listar-entsai", action="store_true",
                   help="Lista os códigos EntSai_es presentes por empresa e sai")
    p.add_argument("--diagnostico-distrato", action="store_true",
                   help="Inspeciona VwDesembolso: colunas, status, datas e contagens")
    # Conexão — sobrescreve .env se passado
    p.add_argument("--server",   default=os.getenv("SQL_SERVER",   "10.30.10.238"))
    p.add_argument("--database", default=os.getenv("SQL_DATABASE", "uau"))
    p.add_argument("--user",     default=os.getenv("SQL_USER",     "ti"))
    p.add_argument("--password", default=os.getenv("SQL_PASSWORD", "ticasaeterra"))
    return p.parse_args()


_ARGS = _parse_args()

SQL_SERVER   = _ARGS.server
SQL_DATABASE = _ARGS.database
SQL_USER     = _ARGS.user
SQL_PASSWORD = _ARGS.password

# ── Parâmetros de execução ────────────────────────────────────────────────────

EMPRESAS           = _ARGS.empresas
DATA_INICIO        = _ARGS.inicio
DATA_FIM           = _ARGS.fim
TIPO_PROC_DISTRATO = _ARGS.tipo_proc

# EntSai_es: 0 = receitas bancárias (rendimento); IR é auto-detectado por empresa
ENTSAI_RECEITA = 0


# ── Conexão ───────────────────────────────────────────────────────────────────

def _detectar_driver() -> str:
    preferencia = [
        "ODBC Driver 18 for SQL Server",
        "ODBC Driver 17 for SQL Server",
        "ODBC Driver 13 for SQL Server",
        "SQL Server Native Client 11.0",
        "SQL Server",
    ]
    disponiveis = pyodbc.drivers()
    for d in preferencia:
        if d in disponiveis:
            return d
    raise RuntimeError(
        f"Nenhum driver ODBC encontrado.\nDisponíveis: {disponiveis}\n"
        "Instale: aka.ms/odbc18"
    )


def conectar() -> pyodbc.Connection:
    driver = _detectar_driver()
    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={SQL_SERVER};"
        f"DATABASE={SQL_DATABASE};"
        f"UID={SQL_USER};"
        f"PWD={SQL_PASSWORD};"
        f"TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str)


# ── Queries ───────────────────────────────────────────────────────────────────

SQL_DISTRATO = """
SELECT
    v.Empresa_Des                       AS Empresa,
    ISNULL(e.Desc_emp, '')              AS Descricao,
    v.Obra_Des                          AS Obra,
    v.DtPgto_Des                        AS Data,
    v.DtGeracao_Des                     AS DtGeracao,
    v.TotalLiq_Des                      AS Valor,
    v.StatusParc_Des                    AS Status,
    NULL                                AS CAP,
    v.Banco_Des                         AS Banco,
    v.ContaCorr_Des                     AS Conta,
    'Distrato'                          AS Tipo,
    v.TipoProc_Des                      AS TipoOrigem,
    v.DescForn_Des                      AS Favorecido
FROM dbo.VwDesembolso v
LEFT JOIN dbo.Empresas e ON v.Empresa_Des = e.Codigo_emp
WHERE v.Empresa_Des = {empresa}
  AND v.TipoProc_Des = {tipo_proc}
  AND v.DtPgto_Des BETWEEN ? AND ?
  {filtro_status}
"""

SQL_RENDIMENTO_IR = """
SELECT
    esa.Empresa_es                                      AS Empresa,
    ISNULL(emp.Desc_emp, '')                            AS Descricao,
    esa.Obra_es                                         AS Obra,
    esa.Data_es                                         AS Data,
    esa.Valor_es                                        AS Valor,
    esa.Cap_es                                          AS CAP,
    esa.Banco_es                                        AS Banco,
    esa.Conta_es                                        AS Conta,
    CASE esa.EntSai_es
        WHEN 0           THEN 'Rendimento Aplicacao'
        WHEN {entsai_ir} THEN 'IR Retido / Despesa Bancaria'
        ELSE CAST(esa.EntSai_es AS VARCHAR)
    END                                                 AS Tipo,
    ISNULL(ctm.Desc_cger, '')                           AS TipoOrigem,
    ISNULL(esa.HistLanc_es, '')                         AS Favorecido
FROM dbo.EntSaiEmpAplic esa
LEFT JOIN dbo.Empresas emp
    ON esa.Empresa_es = emp.Codigo_emp
LEFT JOIN dbo.CategoriasDeTipoDeMovimentacao ctm
    ON esa.Natureza_es = ctm.Codigo_cger
WHERE esa.Empresa_es = {empresa}
  AND esa.Banco_es <> -1
  AND esa.Data_es BETWEEN ? AND ?
  AND (
      esa.EntSai_es = 0
      OR (
          esa.EntSai_es = {entsai_ir}
          AND (
                 UPPER(ISNULL(esa.HistLanc_es,'')) LIKE '%IRRF%'
              OR UPPER(ISNULL(esa.HistLanc_es,'')) LIKE '%I.R.R.F%'
              OR UPPER(ISNULL(esa.HistLanc_es,'')) LIKE '%IMPOSTO DE RENDA%'
              OR UPPER(ISNULL(esa.HistLanc_es,'')) LIKE '%IR RETIDO%'
              OR UPPER(ISNULL(esa.HistLanc_es,'')) LIKE '%I.R. RETIDO%'
              OR UPPER(ISNULL(esa.HistLanc_es,'')) LIKE '% IR'
              OR UPPER(ISNULL(esa.HistLanc_es,'')) LIKE 'IR %'
              OR UPPER(ISNULL(esa.HistLanc_es,'')) LIKE '% IR %'
          )
      )
  )
"""


# ── Busca de dados ────────────────────────────────────────────────────────────
# Condição SQL reutilizada na detecção, validação e na query principal
_IR_KEYWORD_SQL = (
    "UPPER(ISNULL({col},'')) LIKE '%IRRF%' "
    "OR UPPER(ISNULL({col},'')) LIKE '%I.R.R.F%' "
    "OR UPPER(ISNULL({col},'')) LIKE '%IMPOSTO DE RENDA%' "
    "OR UPPER(ISNULL({col},'')) LIKE '%IR RETIDO%' "
    "OR UPPER(ISNULL({col},'')) LIKE '%I.R. RETIDO%' "
    "OR UPPER(ISNULL({col},'')) LIKE '% IR' "
    "OR UPPER(ISNULL({col},'')) LIKE 'IR %' "
    "OR UPPER(ISNULL({col},'')) LIKE '% IR %' "
)


def _ir_cond(col: str = "esa.HistLanc_es") -> str:
    return _IR_KEYWORD_SQL.format(col=col)


def _detectar_entsai_ir(conn: pyodbc.Connection) -> dict[int, int]:
    """Auto-detecta EntSai_es de IR por empresa. Retorna {empresa: cod_ir}."""
    empresas_str = ", ".join(str(e) for e in EMPRESAS)
    cur = conn.cursor()
    cur.execute(
        f"SELECT esa.Empresa_es, emp.Desc_emp, esa.EntSai_es, COUNT(*) AS qtd "
        f"FROM dbo.EntSaiEmpAplic esa "
        f"LEFT JOIN dbo.Empresas emp ON esa.Empresa_es = emp.Codigo_emp "
        f"WHERE esa.Empresa_es IN ({empresas_str}) "
        f"  AND esa.EntSai_es <> 0 "
        f"  AND ({_ir_cond()}) "
        f"GROUP BY esa.Empresa_es, emp.Desc_emp, esa.EntSai_es "
        f"ORDER BY esa.Empresa_es, qtd DESC"
    )
    rows = cur.fetchall()

    # Para cada empresa, pega o código com maior frequência
    por_empresa: dict[int, list[tuple[str, int, int]]] = {}
    for emp_cod, emp_desc, entsai, qtd in rows:
        por_empresa.setdefault(emp_cod, []).append((emp_desc or str(emp_cod), entsai, qtd))

    resultado: dict[int, int] = {}
    print("  🔍 Detectando código IR por empresa...")
    for emp in EMPRESAS:
        candidatos = por_empresa.get(emp, [])
        if candidatos:
            desc, cod, qtd = candidatos[0]  # mais frequente
            resultado[emp] = cod
            extras = ""
            if len(candidatos) > 1:
                outros = ", ".join(f"{c[1]}({c[2]}x)" for c in candidatos[1:])
                extras = f"  outros: {outros}"
            print(f"     Empresa {emp:>4} ({desc[:30]}): EntSai_es = {cod} ({qtd}x com keyword IR){extras}")

            # Alerta se o código é misto (mais registros no período que só os com keyword)
            cur2 = conn.cursor()
            cur2.execute(
                f"SELECT COUNT(*) FROM dbo.EntSaiEmpAplic "
                f"WHERE Empresa_es = {emp} AND EntSai_es = {cod} "
                f"  AND Banco_es <> -1 AND Data_es BETWEEN ? AND ?",
                DATA_INICIO, DATA_FIM,
            )
            total_cod = cur2.fetchone()[0]
            cur2.execute(
                f"SELECT COUNT(*) FROM dbo.EntSaiEmpAplic "
                f"WHERE Empresa_es = {emp} AND EntSai_es = {cod} "
                f"  AND Banco_es <> -1 AND Data_es BETWEEN ? AND ? "
                f"  AND ({_ir_cond('HistLanc_es')})",
                DATA_INICIO, DATA_FIM,
            )
            total_keyword = cur2.fetchone()[0]
            ir_filtrado = total_keyword
            nao_ir = total_cod - total_keyword
            if nao_ir > 0:
                print(f"     ℹ️  Código {cod} é misto no período: {ir_filtrado} IR + {nao_ir} outros (IOF/tarifas) — filtro keyword ativo")
        else:
            resultado[emp] = 44  # fallback
            print(f"     Empresa {emp:>4}: não detectado — usando fallback 44")

    codigos_unicos = set(resultado.values())
    if len(codigos_unicos) > 1:
        print(f"  ⚠️  Códigos IR DIFERENTES entre empresas: {sorted(codigos_unicos)}")
    else:
        print(f"  ✅ Mesmo código IR em todas as empresas: {next(iter(codigos_unicos))}")

    return resultado


def _diagnostico_distrato(conn: pyodbc.Connection) -> None:
    """Inspeciona VwDesembolso para entender por que registros podem estar faltando."""
    cur = conn.cursor()
    empresas_str = ", ".join(str(e) for e in EMPRESAS)

    # 1. Colunas disponíveis na view
    cur.execute("SELECT TOP 0 * FROM dbo.VwDesembolso")
    colunas = [col[0] for col in cur.description]
    print("\n  📌 Colunas em VwDesembolso:")
    for i, c in enumerate(colunas, 1):
        print(f"     {i:>3}. {c}")

    # 2. Contagem total por TipoProc (sem filtro de data)
    cur.execute(
        f"SELECT v.TipoProc_Des, COUNT(*) AS qtd, "
        f"       MIN(v.DtPgto_Des) AS dt_min, MAX(v.DtPgto_Des) AS dt_max "
        f"FROM dbo.VwDesembolso v "
        f"WHERE v.Empresa_Des IN ({empresas_str}) "
        f"GROUP BY v.TipoProc_Des "
        f"ORDER BY qtd DESC"
    )
    print(f"\n  📌 TipoProc presentes (empresa(s) {empresas_str}, sem filtro de data):")
    print(f"  {'TipoProc':>9}  {'Qtd':>6}  {'DtPgto min':>12}  {'DtPgto max':>12}")
    for tp, qtd, dt_min, dt_max in cur.fetchall():
        marca = "  ← filtrado" if tp == TIPO_PROC_DISTRATO else ""
        print(f"  {tp:>9}  {qtd:>6}  {str(dt_min)[:10]:>12}  {str(dt_max)[:10]:>12}{marca}")

    # 3. Registros com DtPgto_Des NULL (pendentes/não pagos) no período por DtEmissao ou DtLanc
    # Descobre qual coluna de data alternativa existe
    cols_data_alt = [c for c in colunas if "Dt" in c and "Pgto" not in c]
    print(f"\n  📌 Outras colunas de data encontradas: {cols_data_alt}")

    cur.execute(
        f"SELECT COUNT(*) "
        f"FROM dbo.VwDesembolso v "
        f"WHERE v.Empresa_Des IN ({empresas_str}) "
        f"  AND v.TipoProc_Des = {TIPO_PROC_DISTRATO} "
        f"  AND v.DtPgto_Des IS NULL"
    )
    n_null = cur.fetchone()[0]
    print(f"\n  📌 Registros TipoProc={TIPO_PROC_DISTRATO} com DtPgto_Des IS NULL: {n_null}")
    if n_null > 0:
        print("     ⚠️  Esses não aparecem no script (filtro por DtPgto_Des). "
              "Se o UAU os exibe, use outra coluna de data.")

    # 4. Contagem com filtro de data atual
    cur.execute(
        f"SELECT COUNT(*) "
        f"FROM dbo.VwDesembolso v "
        f"WHERE v.Empresa_Des IN ({empresas_str}) "
        f"  AND v.TipoProc_Des = {TIPO_PROC_DISTRATO} "
        f"  AND v.DtPgto_Des BETWEEN ? AND ?",
        DATA_INICIO, DATA_FIM,
    )
    n_filtrado = cur.fetchone()[0]
    print(f"  📌 Registros com DtPgto_Des BETWEEN {DATA_INICIO} AND {DATA_FIM}: {n_filtrado}")

    # 5. Coluna de status/situação
    cols_status = [c for c in colunas if any(k in c.lower() for k in ("situa", "status", "situac"))]
    if cols_status:
        for col in cols_status:
            cur.execute(
                f"SELECT v.[{col}], COUNT(*) AS qtd "
                f"FROM dbo.VwDesembolso v "
                f"WHERE v.Empresa_Des IN ({empresas_str}) "
                f"  AND v.TipoProc_Des = {TIPO_PROC_DISTRATO} "
                f"GROUP BY v.[{col}] ORDER BY qtd DESC"
            )
            print(f"\n  📌 Distribuição por [{col}]:")
            for val, qtd in cur.fetchall():
                print(f"     {str(val):<30} {qtd:>6} registros")


def listar_entsai(conn: pyodbc.Connection) -> None:
    """Lista todos os EntSai_es disponíveis por empresa no período — uso: --listar-entsai"""
    cur = conn.cursor()
    empresas_str = ", ".join(str(e) for e in EMPRESAS)
    cur.execute(
        f"SELECT esa.Empresa_es, emp.Desc_emp, esa.EntSai_es, "
        f"       COUNT(*) AS qtd, SUM(esa.Valor_es) AS total "
        f"FROM dbo.EntSaiEmpAplic esa "
        f"LEFT JOIN dbo.Empresas emp ON esa.Empresa_es = emp.Codigo_emp "
        f"WHERE esa.Empresa_es IN ({empresas_str}) "
        f"  AND esa.Banco_es <> -1 "
        f"  AND esa.Data_es BETWEEN ? AND ? "
        f"GROUP BY esa.Empresa_es, emp.Desc_emp, esa.EntSai_es "
        f"ORDER BY esa.Empresa_es, esa.EntSai_es",
        DATA_INICIO, DATA_FIM,
    )
    rows = cur.fetchall()
    if not rows:
        print("  ⚠️  Nenhum registro encontrado para as empresas/período informados.")
        return

    emp_atual = None
    for emp_cod, emp_desc, entsai, qtd, total in rows:
        if emp_cod != emp_atual:
            emp_atual = emp_cod
            print(f"\n  Empresa {emp_cod} — {emp_desc}")
            print(f"  {'Código':>6}  {'Qtd':>6}  {'Total':>15}  Obs")
            print(f"  {'-'*6}  {'-'*6}  {'-'*15}  ---")
        obs = ""
        if entsai == ENTSAI_RECEITA:
            obs = "← rendimento"
        total_fmt = f"{total:,.2f}" if total is not None else "-"
        print(f"  {entsai:>6}  {qtd:>6}  {total_fmt:>15}  {obs}")


def buscar_distrato(conn: pyodbc.Connection, empresa: int) -> pd.DataFrame:
    status_filter = _ARGS.status
    if status_filter is not None:
        placeholders = ", ".join(str(s) for s in status_filter)
        filtro_status = f"AND v.StatusParc_Des IN ({placeholders})"
        print(f"  🔍 [{empresa}] Buscando distratos (TipoProc={TIPO_PROC_DISTRATO}, Status={status_filter})...")
    else:
        filtro_status = ""
        print(f"  🔍 [{empresa}] Buscando distratos (TipoProc={TIPO_PROC_DISTRATO}, todos os status)...")
    sql = SQL_DISTRATO.format(empresa=empresa, tipo_proc=TIPO_PROC_DISTRATO, filtro_status=filtro_status)
    cur = conn.cursor()
    cur.execute(sql, DATA_INICIO, DATA_FIM)
    columns = [col[0] for col in cur.description]
    df = pd.DataFrame.from_records(cur.fetchall(), columns=columns)
    print(f"     → {len(df)} linha(s)")
    return df


def buscar_rendimento_ir(conn: pyodbc.Connection, empresa: int, entsai_ir: int) -> pd.DataFrame:
    print(f"  🔍 [{empresa}] Buscando rendimentos e IR (EntSai_es 0/{entsai_ir})...")
    sql = SQL_RENDIMENTO_IR.format(empresa=empresa, entsai_ir=entsai_ir)
    cur = conn.cursor()
    cur.execute(sql, DATA_INICIO, DATA_FIM)
    columns = [col[0] for col in cur.description]
    df = pd.DataFrame.from_records(cur.fetchall(), columns=columns)
    print(f"     → {len(df)} linha(s)")
    return df


# ── Consolidação ──────────────────────────────────────────────────────────────

COLUNAS_SAIDA = [
    "Empresa", "Descricao", "Obra", "Data", "DtGeracao", "Valor",
    "Status", "CAP", "Banco", "Conta", "Tipo", "TipoOrigem", "Favorecido",
]


def consolidar(df_distrato: pd.DataFrame, df_rendim: pd.DataFrame) -> pd.DataFrame:
    df = pd.concat([df_distrato, df_rendim], ignore_index=True)
    for col in COLUNAS_SAIDA:
        if col not in df.columns:
            df[col] = None
    df = df[COLUNAS_SAIDA]
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    df["Valor"] = (
        pd.to_numeric(df["Valor"], errors="coerce")
        .map(lambda v: f"{v:_.2f}".replace("_", ".").replace(".", "X", 1)
             .replace(".", ",").replace("X", ".") if pd.notna(v) else "")
    )
    df.sort_values(["Empresa", "Data"], inplace=True)
    return df


# ── Exportação XLS ────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
DISTRATO_FILL = PatternFill("solid", fgColor="D9E1F2")
RENDIM_FILL  = PatternFill("solid", fgColor="E2EFDA")
IR_FILL      = PatternFill("solid", fgColor="FCE4D6")


def _cor_linha(tipo: str) -> PatternFill | None:
    if "Distrato" in str(tipo):
        return DISTRATO_FILL
    if "Rendimento" in str(tipo):
        return RENDIM_FILL
    if "IR" in str(tipo):
        return IR_FILL
    return None


def exportar_xls(df: pd.DataFrame, caminho: Path) -> None:
    with pd.ExcelWriter(caminho, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
        df.to_excel(writer, sheet_name="Distrato_Rendimento", index=False)

    wb = load_workbook(caminho)
    ws = wb.active

    # Cabeçalho
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Linhas
    tipo_col_idx = COLUNAS_SAIDA.index("Tipo") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tipo_val = row[tipo_col_idx - 1].value
        fill = _cor_linha(str(tipo_val) if tipo_val else "")
        if fill:
            for cell in row:
                cell.fill = fill

    # Largura automática
    for col_idx, col_name in enumerate(COLUNAS_SAIDA, 1):
        col_letter = get_column_letter(col_idx)
        data_rows = range(2, ws.max_row + 1)
        if data_rows:  # guard: evita crash de max() com sequência vazia
            max_len = max(
                len(str(col_name)),
                *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in data_rows),
            )
        else:
            max_len = len(str(col_name))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    wb.save(caminho)
    print(f"\n✅ Arquivo salvo: {caminho}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    empresas_str = "+".join(str(e) for e in EMPRESAS)
    print(f"\n{'='*60}")
    print(f"  Processando Distrato / Rendimento — Empresa(s): {empresas_str}")
    print(f"  Período: {DATA_INICIO} → {DATA_FIM}")
    print(f"{'='*60}\n")

    conn = conectar()
    try:
        if _ARGS.listar_entsai:
            listar_entsai(conn)
            return

        if _ARGS.diagnostico_distrato:
            _diagnostico_distrato(conn)
            return

        entsai_ir_map = _detectar_entsai_ir(conn)

        frames_distrato = []
        frames_rendim   = []
        for emp in EMPRESAS:
            frames_distrato.append(buscar_distrato(conn, emp))
            frames_rendim.append(buscar_rendimento_ir(conn, emp, entsai_ir_map[emp]))
    finally:
        conn.close()

    df_distrato = pd.concat(frames_distrato, ignore_index=True)
    df_rendim   = pd.concat(frames_rendim,   ignore_index=True)
    df_final    = consolidar(df_distrato, df_rendim)

    if df_final.empty:
        print("\n⚠️  Nenhum dado encontrado para os filtros informados.")
        return

    n_rendim = (df_rendim["Tipo"] == "Rendimento Aplicacao").sum()
    n_ir     = (df_rendim["Tipo"] == "IR Retido / Despesa Bancaria").sum()
    print(f"\n  📊 Total de registros consolidados: {len(df_final)}")
    print(f"     Distratos:        {len(df_distrato)}")
    print(f"     Rendimento Aplic: {n_rendim}")
    print(f"     IR Retido:        {n_ir}")

    nome_arquivo = (
        f"distrato_rendimento_emp{empresas_str}"
        f"_{DATA_INICIO.replace('-', '')}_{DATA_FIM.replace('-', '')}.xlsx"
    )
    destino = _get_base_dir() / nome_arquivo
    exportar_xls(df_final, destino)


if __name__ == "__main__":
    main()

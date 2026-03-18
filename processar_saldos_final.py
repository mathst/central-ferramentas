import pandas as pd
import pyodbc
import sys
import os
import zipfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv


def _get_base_dir() -> Path:
    """Retorna diretorio base: junto ao .exe em bundle, ou CWD em dev.
    Sempre usa sys.executable (onde o .exe está), não _MEIPASS (pasta temp).
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path.cwd()


# Credenciais embarcadas — fallback para .env em dev
_DEFAULTS = {
    "SQL_SERVER":   "10.30.10.238",
    "SQL_DATABASE": "uau",
    "SQL_USER":     "ti",
    "SQL_PASSWORD": "ticasaeterra",
}

load_dotenv(_get_base_dir() / ".env")

SQL_SERVER   = os.getenv("SQL_SERVER",   _DEFAULTS["SQL_SERVER"])
SQL_DATABASE = os.getenv("SQL_DATABASE", _DEFAULTS["SQL_DATABASE"])
SQL_USER     = os.getenv("SQL_USER",     _DEFAULTS["SQL_USER"])
SQL_PASSWORD = os.getenv("SQL_PASSWORD", _DEFAULTS["SQL_PASSWORD"])

EMPRESAS = [19]

# EMPRESAS = [1, 2, 3, 4, 5, 6, 7, 8, 19, 112]
# EMPRESAS = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18,20, 22, 23, 24, 25, 26, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 53, 54, 55, 58, 66, 68, 69, 83, 90, 96, 138]
ANO = 2025


def validar_config() -> None:
    """Valida configuração mínima antes de processar."""
    if not SQL_PASSWORD:
        print("❌ SQL_PASSWORD não definido. Configure a variável de ambiente SQL_PASSWORD.")
        sys.exit(1)
    if not EMPRESAS:
        print("❌ Lista EMPRESAS está vazia.")
        sys.exit(1)
    if ANO < 1900 or ANO > date.today().year + 1:
        print(f"❌ ANO inválido: {ANO}")
        sys.exit(1)


def conectar_banco() -> pyodbc.Connection:
    """Estabelece conexão com SQL Server"""
    try:
        conn_str = (
            f"DRIVER={{ODBC Driver 18 for SQL Server}};"
            f"SERVER={SQL_SERVER};"
            f"DATABASE={SQL_DATABASE};"
            f"UID={SQL_USER};"
            f"PWD={SQL_PASSWORD};"
            f"TrustServerCertificate=yes;"
        )
        conn = pyodbc.connect(conn_str)
        print(f"✅ Conectado: {SQL_DATABASE}@{SQL_SERVER}")
        return conn
    except pyodbc.Error as e:
        print(f"❌ Erro de conexão: {e}")
        sys.exit(1)


def buscar_contas(conn, ano: int) -> pd.DataFrame:
    """Busca contas do PlanoContas para o ano; se vazio, usa o ano mais recente disponível."""
    query = """
    SELECT Conta_plc, Desc_plc, ContaReduz_plc
    FROM PlanoContas
    WHERE Ano_plc = ?
      AND NumMsc_plc = 1
    """
    df = pd.read_sql(query, conn, params=[ano])
    if df.empty:
        query_fallback = """
        SELECT TOP 1 Ano_plc AS ano
        FROM PlanoContas
        WHERE NumMsc_plc = 1
        ORDER BY Ano_plc DESC
        """
        df_ano = pd.read_sql(query_fallback, conn)
        if not df_ano.empty:
            ano_fallback = df_ano.iloc[0]["ano"]
            print(f"   ⚠️  Plano de contas não encontrado para {ano}, usando {ano_fallback}")
            query2 = """
            SELECT Conta_plc, Desc_plc, ContaReduz_plc
            FROM PlanoContas
            WHERE Ano_plc = ?
              AND NumMsc_plc = 1
            """
            df = pd.read_sql(query2, conn, params=[int(ano_fallback)])
    return df


def buscar_movimentos_mensal(conn, empresa: int, ano: int) -> pd.DataFrame:
    """Busca movimentos do ano via LancFiscal+LancSocietario, excluindo lançamentos de abertura (Tipo=2)."""
    query = """
    SELECT Conta_lf AS Conta, Data_lf AS Data, Acao_lf AS Acao, Valor_lf AS Valor
    FROM LancFiscal
    WHERE Empresa_lf = ?
      AND NumMsc_lf = 1
      AND Ano_lf = ?
      AND Tipo_lf <> 2

    UNION ALL

    SELECT Conta_ls, Data_ls, Acao_ls, ValLanc_ls
    FROM LancSocietario
    WHERE Empresa_ls = ?
      AND NumMsc_ls = 1
      AND Ano_ls = ?
      AND Tipo_ls <> 2
    """
    df = pd.read_sql(query, conn, params=[empresa, ano, empresa, ano])
    print(f"   📄 {len(df)} lançamento(s) em {ano} (sem abertura Tipo=2)")
    return df


def buscar_saldo_anterior(conn, empresa: int, ano: int) -> pd.DataFrame:
    """Busca o Saldo Anterior diretamente dos lançamentos de abertura (Tipo=2) do ano atual.
    O UAU grava no 01/01 do ano o saldo transposto do exercício anterior como Tipo=2.
    """
    query = """
    SELECT Conta_lf AS Conta, Acao_lf AS Acao, SUM(Valor_lf) AS Valor
    FROM LancFiscal
    WHERE Empresa_lf = ?
      AND NumMsc_lf = 1
      AND Ano_lf = ?
      AND Tipo_lf = 2
      AND LEFT(Conta_lf, 1) IN ('1', '2')
    GROUP BY Conta_lf, Acao_lf

    UNION ALL

    SELECT Conta_ls, Acao_ls, SUM(ValLanc_ls)
    FROM LancSocietario
    WHERE Empresa_ls = ?
      AND NumMsc_ls = 1
      AND Ano_ls = ?
      AND Tipo_ls = 2
      AND LEFT(Conta_ls, 1) IN ('1', '2')
    GROUP BY Conta_ls, Acao_ls
    """
    df = pd.read_sql(query, conn, params=[empresa, ano, empresa, ano])
    print(f"   📄 {len(df)} linha(s) de abertura Tipo=2 (Saldo Anterior)")
    return df


def _sort_key(conta: str):
    """Ordena contas hierarquicamente: 1 < 1.01 < 1.01.01 < 2 < ..."""
    try:
        return tuple(int(p) for p in str(conta).split("."))
    except ValueError:
        return (999999,)


def calcular_balancete(
    df_contas: pd.DataFrame,
    df_lanc_ant: pd.DataFrame,
    df_lanc: pd.DataFrame,
    ano: int,
) -> pd.DataFrame:
    # Só aborta se não há movimentos no ano E não há histórico anterior
    if df_lanc.empty and df_lanc_ant.empty:
        return pd.DataFrame()

    # --- Metadados das contas (sem Anexos, ResFiscal, ContaRef) ---
    contas_info = df_contas.rename(columns={
        "Conta_plc": "Conta",
        "Desc_plc": "Descricao",
        "ContaReduz_plc": "ContaReduz",
    }).copy()
    contas_info["eh_devedora"] = contas_info["Conta"].astype(str).str[0].isin(["1", "4"])
    conta_nat = contas_info.set_index("Conta")["eh_devedora"]

    valid_contas = set(contas_info["Conta"].astype(str))

    def get_ancestors(conta: str) -> list:
        parts = str(conta).split(".")
        return [".".join(parts[:i]) for i in range(1, len(parts) + 1)]

    # --- Saldo Anterior (via lançamentos Tipo=2 do ano atual) ---
    # O UAU transpõe o saldo do exercício anterior como lançamentos Tipo=2 em 01/01.
    # df_lanc_ant já vem agrupado por Conta+Acao com a soma dos valores.
    saldo_ant_map: dict = {}
    if not df_lanc_ant.empty:
        dfa = df_lanc_ant[df_lanc_ant["Conta"].astype(str).isin(valid_contas)].copy()
        dfa["Valor"] = pd.to_numeric(dfa["Valor"], errors="coerce").fillna(0)
        dfa["Acao"] = dfa["Acao"].astype(str).str.upper().str.strip()
        dfa["eh_dev"] = dfa["Conta"].map(conta_nat).fillna(True)
        positivo = (dfa["eh_dev"] & (dfa["Acao"] == "D")) | (~dfa["eh_dev"] & (dfa["Acao"] == "C"))
        dfa["Movimento"] = dfa["Valor"].where(positivo, -dfa["Valor"])
        dfa["Contas_all"] = dfa["Conta"].apply(get_ancestors)
        dfa_exp = dfa.explode("Contas_all")
        saldo_ant_map = {
            k: Decimal(str(v))
            for k, v in dfa_exp.groupby("Contas_all")["Movimento"].sum().items()
        }

    # --- Movimentos do ano corrente ---
    if not df_lanc.empty:
        df = df_lanc[df_lanc["Conta"].astype(str).isin(valid_contas)].copy()
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
        df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce").fillna(0)
        df["Acao"] = df["Acao"].astype(str).str.upper().str.strip()
        df["Mes"] = df["Data"].dt.to_period("M")
        df["Contas_all"] = df["Conta"].apply(get_ancestors)

        df_exp = df.explode("Contas_all").rename(columns={"Conta": "Conta_orig", "Contas_all": "Conta"})

        df_deb = (
            df_exp[df_exp["Acao"] == "D"]
            .groupby(["Conta", "Mes"])["Valor"]
            .sum()
            .rename("Debito")
        )
        df_cred = (
            df_exp[df_exp["Acao"] == "C"]
            .groupby(["Conta", "Mes"])["Valor"]
            .sum()
            .rename("Credito")
        )
        df_mensal = pd.concat([df_deb, df_cred], axis=1).fillna(0).reset_index()
        # Garante que não há duplicatas (Conta, Mes) após o concat de LancFiscal+LancSocietario
        df_mensal = df_mensal.groupby(["Conta", "Mes"], as_index=False)[["Debito", "Credito"]].sum()
        contas_com_mov = set(df_mensal["Conta"].unique())
        meses_com_mov = sorted(df_mensal["Mes"].unique())
    else:
        df_mensal = pd.DataFrame(columns=["Conta", "Mes", "Debito", "Credito"])
        contas_com_mov = set()
        meses_com_mov = []

    # Normaliza tipo da coluna Conta para string em ambos os DataFrames
    df_mensal["Conta"] = df_mensal["Conta"].astype(str)
    contas_info["Conta"] = contas_info["Conta"].astype(str)
    conta_nat.index = conta_nat.index.astype(str)

    result_rows = []
    for _, row in contas_info.iterrows():
        conta = str(row["Conta"])
        conta_data = df_mensal[df_mensal["Conta"] == conta].set_index("Mes")
        eh_dev = row["eh_devedora"]
        saldo_running = Decimal(str(saldo_ant_map.get(conta, 0.0)))

        record: dict = {
            "Conta Reduz.": row["ContaReduz"],
            "Conta": conta,
            "Descrição": row["Descricao"],
            "Saldo Anterior": float(saldo_running),
        }

        for mes in meses_com_mov:
            deb = Decimal(str(conta_data.at[mes, "Debito"])) if mes in conta_data.index else Decimal(0)
            cred = Decimal(str(conta_data.at[mes, "Credito"])) if mes in conta_data.index else Decimal(0)
            saldo_running = saldo_running + deb - cred if eh_dev else saldo_running - deb + cred
            record[f"Saldo {mes.strftime('%m/%Y')}"] = float(saldo_running)

        todos_saldos = [record["Saldo Anterior"]] + [
            record[f"Saldo {mes.strftime('%m/%Y')}"] for mes in meses_com_mov
        ]
        if any(s != 0.0 for s in todos_saldos):
            result_rows.append(record)

    result = pd.DataFrame(result_rows)

    if result.empty:
        print(f"   📊 0 conta(s) com valor | {len(meses_com_mov)} mês(es) lançados")
        return result

    result["_sort"] = result["Conta"].apply(_sort_key)
    result = result.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)

    print(f"   📊 {len(result)} conta(s) com valor | {len(meses_com_mov)} mês(es) lançados")
    return result


def processar_empresa(conn, empresa: int, ano: int) -> pd.DataFrame:
    print(f"\n  🏢 Empresa {empresa}")
    df_contas = buscar_contas(conn, ano)
    print(f"   📋 {len(df_contas)} conta(s) no plano de contas")
    df_lanc_ant = buscar_saldo_anterior(conn, empresa, ano)
    df_lanc = buscar_movimentos_mensal(conn, empresa, ano)
    return calcular_balancete(df_contas, df_lanc_ant, df_lanc, ano)


def salvar_excel_empresa(empresa: int, df: pd.DataFrame, ano: int):
    arquivo = f"BALANCETE_{ano}_EMP_{empresa}.xlsx"
    print(f"\n💾 Salvando: {arquivo}")

    fmt_brl = '#,##0.00'

    if df.empty:
        print(f"   ⚠️  Empresa {empresa}: sem dados, arquivo não gerado.")
        return

    df = df.copy()
    idx_mon = {
        i for i, c in enumerate(df.columns)
        if c == "Saldo Anterior" or c.startswith("Saldo ")
    }
    for i in idx_mon:
        df.iloc[:, i] = pd.to_numeric(df.iloc[:, i], errors="coerce")

    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
        wb = writer.book
        ws = wb.create_sheet(title="Balancete")
        # remove a aba vazia padrão do openpyxl
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        writer.sheets["Balancete"] = ws

        for col_i, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_i, value=col_name)

        for row_i, tup in enumerate(df.itertuples(index=False), 2):
            for col_i, val in enumerate(tup, 1):
                if isinstance(val, float) and pd.isna(val):
                    val = None
                cell = ws.cell(row=row_i, column=col_i, value=val)
                if (col_i - 1) in idx_mon and val is not None:
                    cell.number_format = fmt_brl

        for i, col in enumerate(df.columns, 1):
            max_len = max(df.iloc[:, i-1].astype(str).apply(len).max(), len(str(col))) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max_len, 45)

        ws.freeze_panes = "A2"

    print(f"   ✅ {len(df)} conta(s) → {arquivo}")


def main():
    validar_config()

    print("=" * 70)
    print(f"  BALANCETE — Empresas: {EMPRESAS} | Ano: {ANO}")
    print("=" * 70)

    conn = conectar_banco()
    arquivos_gerados = []
    try:
        for empresa in EMPRESAS:
            df = processar_empresa(conn, int(empresa), ANO)
            if df.empty:
                print(f"\n⚠️  Empresa {empresa}: sem dados, arquivo não gerado.")
                continue
            salvar_excel_empresa(int(empresa), df, ANO)
            arquivos_gerados.append(f"BALANCETE_{ANO}_EMP_{int(empresa)}.xlsx")
    finally:
        conn.close()

    if len(arquivos_gerados) > 4:
        zip_nome = f"BALANCETE_{ANO}.zip"
        with zipfile.ZipFile(zip_nome, "w", zipfile.ZIP_DEFLATED) as zf:
            for arq in arquivos_gerados:
                zf.write(arq)
        print(f"\n📦 {len(arquivos_gerados)} arquivos compactados → {zip_nome}")

    print(f"\n🔌 Conexão encerrada")
    print(f"\n✅ Concluído!")


if __name__ == "__main__":
    main()